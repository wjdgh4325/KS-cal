import torch
import util
# import os
# os.environ['CUDA_LAUNCH_BLOCKING'] = '1'
from args import TestArgParser
from saver import ModelSaver
import util
import numpy as np
from SurvivalEVAL.Evaluator import QuantileRegEvaluator
from SurvivalEVAL.utils.util_survival import compute_decensor_times, make_mono_quantiles
from SurvivalEVAL.utils.cox_cdf import pred_params_to_cox
import pandas as pd
from scipy.interpolate import interp1d, PchipInterpolator
from typing import List, Tuple, Union
from openpyxl import load_workbook
import time

DEVICE = torch.device("cuda" if torch.cuda.is_available() else "cpu")

def make_monotonic(
        survival_curves: np.ndarray,
        times_coordinate: np.ndarray,
        method: str = "ceil",
        seed: int = None,
        num_bs: int = None
):
    """
    Make the survival curves monotonic.
    Parameters
    ----------
    survival_curves: np.ndarray
        Survival curves. 2-D array of survival probabilities. The first dimension is the number of samples. The second
        dimension is the number of time points.
    times_coordinate: np.ndarray
        Time points corresponding to the survival curves. 1-D array of time points.
    method: str
        The method to make the survival curves monotonic. One of ['ceil', 'floor', 'bootstrap']. Default: 'ceil'.
    seed: int
        Random seed for bootstrapping. Default: None.
    num_bs: int
        Number of bootstrap samples. Default: None. If None, then num_bs = 10 * num_times.

    Returns
    -------
    survival_curves: np.ndarray
        Survival curves with monotonicity. 2-D array of survival probabilities.
    """
    # if np.all(np.sort(times_coordinate) != times_coordinate):
    #     raise ValueError("The time coordinates must be sorted in ascending order.")

    if num_bs is None:
        # 10 times the number of time points or 1000, whichever is larger
        num_bs = max(10 * len(times_coordinate), 1000)

    if seed is not None:
        np.random.seed(seed)

    survival_curves = np.clip(survival_curves, 0, 1)
    if not check_monotonicity(survival_curves):
        if method == "ceil":
            survival_curves = np.maximum.accumulate(survival_curves[:, ::-1], axis=1)[:, ::-1]
        elif method == "floor":
            survival_curves = np.minimum.accumulate(survival_curves, axis=1)
        elif method == "bootstrap":
            need_rearrange = np.where(np.any((np.sort(survival_curves, axis=1)[:, ::-1] != survival_curves), axis=1))[0]

            for i in need_rearrange:
                inter_lin = interp1d(survival_curves[i], times_coordinate, kind='linear', fill_value='extrapolate')
                # Bootstrap the quantile function
                bootstrap_qf = inter_lin(np.random.uniform(0, 1, num_bs))
                # Now compute the rearranged survival curve
                # The original method is to compute a value (time) given the fixed quantile (probability)
                # Here we compute the probability (quantile) given the fixed value (time)
                for j, time in enumerate(times_coordinate):
                    survival_curves[i, j] = np.mean(bootstrap_qf > time)
        else:
            raise ValueError("method must be one of ['ceil', 'floor', 'bootstrap']")
    
    small_value = 1e-10
    survival_curves -= np.cumsum(np.full(survival_curves.shape, small_value), axis=1)

    return np.clip(survival_curves, 0, 1)

Numeric = Union[float, int, bool]
NumericArrayLike = Union[List[Numeric], Tuple[Numeric], np.ndarray, pd.Series, pd.DataFrame, torch.Tensor]

def check_and_convert(*args):
    """ Makes sure that the given inputs are numpy arrays, list,
        tuple, panda Series, pandas DataFrames, or torch Tensors.

        Also makes sure that the given inputs have the same shape.

        Then convert the inputs to numpy array.

        Parameters
        ----------
        * args : tuple of objects
                 Input object to check / convert.

        Returns
        -------
        * result : tuple of numpy arrays
                   The converted and validated arg.

        If the input isn't numpy arrays, list or pandas DataFrames, it will
        fail and ask to provide the valid format.
    """

    result = ()
    last_length = ()
    for i, arg in enumerate(args):

        if len(arg) == 0:
            error = " The input is empty. "
            error += "Please provide at least 1 element in the array."
            raise IndexError(error)

        else:

            if isinstance(arg, np.ndarray):
                x = (arg.astype(np.double),)
            elif isinstance(arg, list):
                x = (np.asarray(arg).astype(np.double),)
            elif isinstance(arg, tuple):
                x = (np.asarray(arg).astype(np.double),)
            elif isinstance(arg, pd.Series):
                x = (arg.values.astype(np.double),)
            elif isinstance(arg, pd.DataFrame):
                x = (arg.values.astype(np.double),)
            elif isinstance(arg, torch.Tensor):
                x = (arg.cpu().numpy().astype(np.double),)
            else:
                error = """{arg} is not a valid data format. Only use 'list', 'tuple', 'np.ndarray', 'torch.Tensor',
                        'pd.Series', 'pd.DataFrame'""".format(arg=type(arg))
                raise TypeError(error)

            if np.sum(np.isnan(x)) > 0.:
                error = "The #{} argument contains null values"
                error = error.format(i + 1)
                raise ValueError(error)

            if len(args) > 1:
                if i > 0:
                    assert x[0].shape == last_length, """Shapes between {}-th input array and
                    {}-th input array are not consistent""".format(i - 1, i)
                result += x
                last_length = x[0].shape
            else:
                result = x[0]

    return result

def check_monotonicity(array: NumericArrayLike):
    array = check_and_convert(array)
    if array.ndim == 1:
        return (all(array[i] <= array[i + 1] for i in range(len(array) - 1)) or
                all(array[i] >= array[i + 1] for i in range(len(array) - 1)))
    elif array.ndim == 2:
        return (all(all(array[:, i] <= array[:, i + 1]) for i in range(array.shape[1] - 1)) or
                all(all(array[:, i] >= array[:, i + 1]) for i in range(array.shape[1] - 1)))
    else:
        raise ValueError("The input array must be 1-D or 2-D.")

def survival_to_quantile(surv_prob, time_coordinates, quantile_levels, interpolate='Pchip'):
    if interpolate == 'Linear':
        Interpolator = interp1d
    elif interpolate == 'Pchip':
        Interpolator = PchipInterpolator
    else:
        raise ValueError(f"Unknown interpolation method: {interpolate}")
    cdf = 1 - surv_prob
    slope = cdf[:, -1] / time_coordinates[:, -1]
    assert cdf.shape == time_coordinates.shape, "CDF and time coordinates have different shapes."
    quantile_predictions = np.empty((cdf.shape[0], quantile_levels.shape[0]))
    for i in range(cdf.shape[0]):
        # fit a scipy interpolation function to the cdf
        cdf_i = cdf[i, :]
        time_coordinates_i = time_coordinates[i, :]
        # remove duplicates in cdf_i (x-axis), otherwise Interpolator will raise an error
        # here we only keep the first occurrence of each unique value
        cdf_i, idx = np.unique(cdf_i, return_index=True)
        time_coordinates_i = time_coordinates_i[idx]
        interp = Interpolator(cdf_i, time_coordinates_i)

        # if the quantile level is beyond last cdf, we extrapolate the
        beyond_last_idx = np.where(quantile_levels > cdf_i[-1])[0]
        quantile_predictions[i] = interp(quantile_levels)
        quantile_predictions[i, beyond_last_idx] = quantile_levels[beyond_last_idx] / slope[i]

    # sanity checks
    assert np.all(quantile_predictions >= 0), "Quantile predictions contain negative."
    assert check_monotonicity(quantile_predictions), "Quantile predictions are not monotonic."
    return quantile_predictions

def interpolated_survival_curve(times_coordinate, survival_curve, interpolation):
    if interpolation == "Linear":
        spline = interp1d(times_coordinate, survival_curve, kind='linear', fill_value='extrapolate')
    elif interpolation == "Pchip":
        spline = PchipInterpolator(times_coordinate, survival_curve)
    else:
        raise ValueError("interpolation must be one of ['Linear', 'Pchip']")
    return spline

def quantile_to_survival(quantile_levels, quantile_predictions, time_coordinates, interpolate='Pchip'):
    survival_level = 1 - quantile_levels
    slope = - quantile_levels[-1] / quantile_predictions[:, -1]
    surv_pred = np.empty((quantile_predictions.shape[0], time_coordinates.shape[0]))
    for i in range(quantile_predictions.shape[0]):
        # fit an interpolation function to the cdf
        spline = interpolated_survival_curve(quantile_predictions[i, :], survival_level, interpolate)

        # if the quantile level is beyond last cdf, we extrapolate the
        beyond_prob_idx = np.where(time_coordinates > quantile_predictions[i, -1])[0]
        surv_pred[i] = spline(time_coordinates)
        surv_pred[i, beyond_prob_idx] = np.clip(time_coordinates[beyond_prob_idx] * slope[i] + 1,
                                                a_min=0, a_max=1)
    surv_pred = make_monotonic(surv_pred, times_coordinate=time_coordinates)
    print(surv_pred)
    # sanity checks
    assert np.all(surv_pred >= 0), "Survival predictions contain negative."
    assert check_monotonicity(surv_pred), "Survival predictions are not monotonic."
    return surv_pred

def xcal_from_hist(d_cal_hist: np.ndarray):
    """
    Compute the x-calibration score from the D-calibration histogram.
    """
    # get bin number
    n_bins = d_cal_hist.shape[0]
    # normalize the histogram
    d_cal_hist = d_cal_hist / d_cal_hist.sum()
    # compute the x-calibration score
    optimal = np.ones_like(d_cal_hist) / n_bins
    # 1/(n_bins-1) is because there is only (n_bins-1) degrees of freedom for n_bins
    x_cal = (1 / (n_bins - 1)) * np.sum(np.square(d_cal_hist.cumsum() - optimal.cumsum()))
    return x_cal

def make_strictly_increasing_below_one(arr, eps=1e-8):
    arr = np.array(arr, dtype=np.float64)
    for i in range(1, len(arr)):
        if arr[i] <= arr[i - 1]:
            arr[i] = arr[i - 1] + eps
        if arr[i] >= 1.0:
            arr[i] = 1.0 - (len(arr) - i) * eps  # 뒤로 갈수록 더 작게
    return arr

def CSD(args):
    model, ckpt_info = ModelSaver.load_model(args.ckpt_path, args)
    args.start_epoch = ckpt_info['epoch'] + 1
    args.device = DEVICE
    model = model.to(args.device)
    model.eval()

    train_loader = util.get_train_loader(args, during_training=False)
    eval_loaders = util.get_eval_loaders(during_training=False, args=args)
    valid_loader, test_loader = eval_loaders

    args.loss_fn = 'mle'

    for src_train, tgt_train in train_loader:
        src_train = src_train
        tte_train = tgt_train[:, 0]
        is_dead_train = tgt_train[:, 1]
        order_train = torch.argsort(tte_train)

    for src_valid, tgt_valid in valid_loader:
        src_valid = src_valid.to(DEVICE)
        tgt_valid = tgt_valid.to(DEVICE)
        tte_valid = tgt_valid[:, 0]
        is_dead_valid = tgt_valid[:, 1]
        order_valid = torch.argsort(tte_valid)

    for src_test, tgt_test in test_loader:
        src_test = src_test.to(DEVICE)
        tgt_test = tgt_test.to(DEVICE)
        tte_test = tgt_test[:, 0]
        is_dead_test = tgt_test[:, 1]
        order_test = torch.argsort(tte_test)

    pred_params_valid = model.forward(src_valid)
    pred_params_test = model.forward(src_test)
    if args.model_dist in ['cat', 'mtlr']:
        tgt_valid = util.cat_bin_target(args, tgt_valid, args.bin_boundaries)
        tgt_test = util.cat_bin_target(args, tgt_test, args.bin_boundaries)
        
    cdf_valid_all = util.get_cdf_matrix(pred_params_valid, tgt_valid, args)
    cdf_test_all = util.get_cdf_matrix(pred_params_test, tgt_test, args)

    cdf_valid = util.get_cdf_val(pred_params_valid, tgt_valid, args)

    if args.model_dist != 'cox':
        cdf_valid_all = cdf_valid_all[:, order_valid][order_valid, :]
        cdf_test_all = cdf_test_all[:, order_test][order_test, :]
        cdf_valid = cdf_valid[order_valid]

    else:
        cdf_valid_all = cdf_valid_all
        cdf_test_all = cdf_test_all
        cdf_valid = cdf_valid
        
    tte_valid = tte_valid[order_valid]
    is_dead_valid = is_dead_valid[order_valid]

    tte_test = tte_test[order_test]
    is_dead_test = is_dead_test[order_test]

    start_time = time.time()
    n_quantiles = 9
    cdf_valid_all = torch.clamp(cdf_valid_all, min=0, max=1)
    cdf_test_all = torch.clamp(cdf_test_all, min=0, max=1)
    surv_valid = 1 - cdf_valid_all
    surv_test = 1 - cdf_test_all
    surv_valid = make_monotonic(surv_valid.cpu().numpy(), tte_valid.cpu().numpy())
    surv_test = make_monotonic(surv_test.cpu().numpy(), tte_test.cpu().numpy())
    print(check_monotonicity(surv_valid))
    print(check_monotonicity(surv_test))
    
    csd_quantile = np.linspace(1 / (n_quantiles + 1), n_quantiles / (n_quantiles + 1), n_quantiles)
    
    if tte_test[0] != 0:
        tte_test2 = np.concatenate([np.array([0]), tte_test.cpu().numpy()], 0)
        surv_test2 = np.concatenate([np.ones([surv_test.shape[0], 1]), surv_test], 1)

    tte_test2 = np.repeat(tte_test2[np.newaxis, :], surv_test2.shape[0], axis=0)
    quantile_prediction_test = survival_to_quantile(surv_test2, tte_test2, quantile_levels=csd_quantile)

    R = 1000
    grid = np.linspace(0, 1, R).reshape(R, 1)
    surv_con = 1 - cdf_valid
    con_scores_event = np.repeat(surv_con[is_dead_valid == 1].cpu().numpy(), R)
    con_scores_cen = grid * surv_con[is_dead_valid == 0].cpu().numpy().reshape(1, -1)
    con_scores_cen = con_scores_cen.reshape(-1)
    conformal_score = np.concatenate([con_scores_event, con_scores_cen])
    conformal_score = np.sort(conformal_score, axis=0)
    index = np.ceil((1-csd_quantile) * (conformal_score.shape[0]+1)) - 1
    index = index.astype(int)
    index = index.clip(0, conformal_score.shape[0] - 1)
    adjusted_quantile = 1 - conformal_score[index]
    adjusted_quantile = make_strictly_increasing_below_one(adjusted_quantile)
    quantile_prediction_test2 = survival_to_quantile(surv_test2, tte_test2, quantile_levels=adjusted_quantile)
    print("Adjusted quantiles:", adjusted_quantile)
    csd_quantile, quantile_prediction_test2 = make_mono_quantiles(csd_quantile, quantile_prediction_test2, method='bootstrap')
    if 0 not in csd_quantile:
        csd_quantile = np.insert(csd_quantile, 0, 0)
        quantile_prediction_test2 = np.insert(quantile_prediction_test2, 0, 0)
    assert np.all(quantile_prediction_test2 >= 0), "Quantile predictions contain negative."
    assert check_monotonicity(quantile_prediction_test2), "Quantile predictions are not monotonic."
    print("Quantile:", quantile_prediction_test2)
    # t_train_val = np.concatenate((tte_train, tte_valid.cpu()))
    # e_train_val = np.concatenate((is_dead_train, is_dead_valid.cpu()))
    end_time = time.time()

    evaler = QuantileRegEvaluator(quantile_prediction_test2, csd_quantile, tte_test, is_dead_test, tte_train, is_dead_train,
                                    predict_time_method="Mean", interpolation='Pchip')
    surv_test = evaler.predict_probability_from_curve(tte_test.cpu().numpy())
    C_index = evaler.concordance()[0]

    S_cal = util.s_calibration(points=torch.tensor(1-surv_test), phase=args.phase, is_dead=is_dead_test.cpu(), args=args)
    D_cal = util.d_calibration(points=torch.tensor(1-surv_test), is_dead=is_dead_test.cpu(), args=args, phase=args.phase)
    KS_cal, KS = util.get_p_value(args=args, cdf=torch.tensor(1-surv_test), is_dead=is_dead_test.cpu())

    quan_to_surv = quantile_to_survival(quantile_levels=csd_quantile, quantile_predictions=quantile_prediction_test2,
                                        time_coordinates=tte_test.cpu().numpy())
    KM_cal = util.km_calibration(cdf=torch.tensor(1-quan_to_surv).to(DEVICE), tte=tte_test, is_dead=is_dead_test, device=DEVICE)
    IBS = util.integrated_brier_score(train_tte=tte_train, train_event=is_dead_train, test_tte=tte_test, test_event=is_dead_test,
                                      cdf_test=torch.tensor(1-quan_to_surv).to(DEVICE), time=tte_test)
    print("C-index:", C_index)
    print("S-cal:", S_cal.item())
    print("D-cal:", D_cal.item())
    print("KS metric:", KS.item())
    print("KM-cal:", KM_cal.item())
    print("IBS:", IBS.item())
    print("CSD-iPOT time:", end_time - start_time)
    # workbook = load_workbook(filename='./tmp2.xlsx')
    # sheet = workbook.active
    # last_row = sheet.max_row
    # sheet.cell(row=last_row+1, column=1, value=(end_time - start_time))
    # sheet.cell(row=last_row+1, column=2, value=(str(args.dataset) + '_' + str(args.model_dist)))
    # workbook.save('./tmp2.xlsx')

    workbook = load_workbook(filename='./tmp2.xlsx')
    sheet = workbook.active
    last_row = sheet.max_row
    sheet.cell(row=last_row+1, column=1, value=C_index)
    sheet.cell(row=last_row+1, column=2, value=S_cal.item())
    sheet.cell(row=last_row+1, column=3, value=D_cal.item())
    sheet.cell(row=last_row+1, column=4, value=KS.item())
    # sheet.cell(row=last_row+1, column=5, value=KS_cal.item())
    sheet.cell(row=last_row+1, column=5, value=KM_cal.item())
    sheet.cell(row=last_row+1, column=6, value=IBS.item())
    sheet.cell(row=last_row+1, column=7, value=(f'CSD-iPOT_{args.dataset}_{args.model_dist}_lam_{args.lam}'))
    workbook.save('./tmp2.xlsx')
    # ci = []
    # mae_hinge = []
    # mae_po = []
    # rmse_hinge = []
    # rmse_po = []
    # ibs = []
    # km_cal = []
    # xcal_stats = []

    # c_index = evaler.concordance()[0]
    # ibs_score = evaler.integrated_brier_score(num_points=10)
    # hinge_abs = evaler.mae(method='Hinge', verbose=False)
    # po_abs = evaler.mae(method='Pseudo_obs', verbose=False)
    # hinge_sq = evaler.rmse(method='Hinge', verbose=False)
    # po_sq = evaler.rmse(method='Pseudo_obs', verbose=False)
    # km_cal_score = evaler.km_calibration()
    # _, dcal_hist = evaler.d_calibration()
    # xcal_score = xcal_from_hist(dcal_hist)

    # ci.append(c_index)
    # ibs.append(ibs_score)
    # mae_hinge.append(hinge_abs)
    # mae_po.append(po_abs)
    # rmse_hinge.append(hinge_sq)
    # rmse_po.append(po_sq)
    # km_cal.append(km_cal_score)
    # xcal_stats.append(xcal_score)

    # print("concordance:", ci)
    # print("IBS:", ibs)
    # print("MAE_hinge:", mae_hinge)
    # print("MAE_po:", mae_po)
    # print("KM-CAL:", km_cal)
    # print("X-cal:", xcal_stats)


if __name__ == '__main__':
    parser = TestArgParser()
    args = parser.parse_args()

    if args.dataset == 'mnist':
        assert args.model == 'SurvMNISTNN', "if dataset == mnist, model must be SurvMNISTNN"

    if args.model_dist in ['cat', 'mtlr']:
        bin_boundaries, mid_points = util.get_bin_boundaries(args)
        args.bin_boundaries = bin_boundaries
        args.mid_points = mid_points
        # args.marginal_counts = marginal_counts

    with torch.no_grad():
        metrics = CSD(args)
